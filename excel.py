import openpyxl as xl
import datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

class Record:

    def __init__(
            self,
            date: datetime.date|str,
            description: str,
            account: str,
            transaction_type: str,
            supply_value: int|str,
            vat: int|str,
            remark: str
        ):
        """
        date: 'YY/mm/dd' 또는 date 객체
        supply_value: '0,000,000' 또는 int
        vat: '0,000,000' 또는 int
        """
        self.__date = __class__.convert_date(date)
        self.__description = description
        self.__account = account
        self.__transaction_type = transaction_type
        self.__supply_value = __class__.convert_int(supply_value)
        self.__vat = __class__.convert_int(vat)
        self.__remark = remark
    
    @property
    def date(self): return self.__date
    @property
    def description(self): return self.__description
    @property
    def account(self): return self.__account
    @property
    def transaction_type(self): return self.__transaction_type
    @property
    def supply_value(self): return self.__supply_value
    @property
    def vat(self): return self.__vat
    @property
    def remark(self): return self.__remark

    @property
    def desc(self): return self.description
    @property
    def acc(self): return self.account
    @property
    def tt(self): return self.transaction_type
    @property
    def sv(self): return self.supply_value
    
    @classmethod
    def convert_date(cls, date:datetime.date|str) -> datetime.date:
        if isinstance(date, str):
            year, month, day = date.split('/')
            year = int(f'20{year}')
            month = int(month)
            day = int(day)
            return datetime.date(year, month, day)
        return date

    @classmethod
    def convert_int(cls, number:int|str) -> int:
        if isinstance(number, str):
            return int(number.replace(',', ''))
        return number

def save_records_as_excel(excel_file_path:str, records:list[Record,]):
    font_bold = Font(bold=True)
    fill_header = PatternFill(fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')
    border = Border(
        left   = Side(style='thin'), 
        right  = Side(style='thin'), 
        top    = Side(style='thin'), 
        bottom = Side(style='thin')
    )
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = xl.Workbook()
    ws = wb.active
    ws['A1'].value = '일자'
    ws['B1'].value = '거래내용'
    ws['C1'].value = '거래처'
    ws['D1'].value = '수입'
    ws['D2'].value = '금액'
    ws['E2'].value = '부가세'
    ws['F1'].value = '비용'
    ws['F2'].value = '금액'
    ws['G2'].value = '부가세'
    ws['H1'].value = '고정자산'
    ws['H2'].value = '금액'
    ws['I2'].value = '부가세'
    ws['J1'].value = '비고'
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')
    ws.merge_cells('H1:I1')
    ws.merge_cells('J1:J2')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 40
    for i in range(1, 3):
        for j in range(1, 11):
            cell = ws.cell(row=i, column=j)
            cell.fill = fill_header
            cell.font = font_bold
            cell.border = border
            cell.alignment = alignment_center
    for i, record in enumerate(records):
        i_row = i+3
        if record.tt == '수입':
            offset = 0
        elif record.tt == '비용':
            offset = 2
        elif record.tt == '고정자산':
            offset = 4
        else:
            raise ValueError(f'Invalid transaction type: {record.tt}')
        ws.cell(row=i_row, column=1).value = record.date
        ws.cell(row=i_row, column=2).value = record.desc
        ws.cell(row=i_row, column=3).value = record.acc
        ws.cell(row=i_row, column=4+offset).value = record.sv
        ws.cell(row=i_row, column=5+offset).value = record.vat
        ws.cell(row=i_row, column=10).value = record.remark
        for j in range(1, 11):
            cell = ws.cell(row=i_row, column=j)
            cell.border = border
            cell.alignment = alignment_center
            if 4 <= j <= 9:
                cell.number_format = '#,###'

    wb.save(excel_file_path)
    wb.close()